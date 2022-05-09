import openpyxl

class toExcel:
    
    @classmethod
    def cargarBrief(cls,comando):    
        x= range(0,len(comando))
        contador=3
        for todos in x:
            book =openpyxl.load_workbook('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
            contador=contador+1
            sheet = book['interfaces']
            sheet[f'A{contador}']=comando[todos]
            book.save('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
        print('FIN BRIEF') 
        
        
        
    @classmethod    
    def cargarRun(cls,comando):    
        x= range(0,len(comando))
        contador=3
        for todos in x:
            book =openpyxl.load_workbook('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
            contador=contador+1
            sheet = book['configuracion']
            sheet[f'A{contador}']=comando[todos]
            book.save('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
        print('FIN SH RUN') 
        
    @classmethod    
    def cargarVersion(cls,comando):    
        x= range(0,len(comando))
        contador=3
        for todos in x:
            contador=contador+1
            book =openpyxl.load_workbook('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
            sheet = book['configuracion']
            sheet[f'I{contador}']=comando[todos]
            book.save('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
        print('FIN VERSION')     

    @classmethod    
    def cargarEncabezado(cls,lista):    
            book =openpyxl.load_workbook('/home/fr/Documentos/pythonEntel/servicio/store/datos.xlsx')
            sheet = book['encabezado']
            sheet['B3']=lista[0]
            sheet['B4']=lista[1]
            sheet['B11']=lista[2]
            sheet['B17']=lista[3]
            sheet['B22']=lista[4]
            sheet['B19']=lista[5]
            sheet['B23']=lista[6]
            book.save(f'/home/fr/Documentos/pythonEntel/servicio/store/check/CHECKLIST_ITP_{lista[0]}_{lista[1]}_{lista[2]}.xlsx')
            print('FIN ENCABEZADO')   