
from ping3 import ping 
import openpyxl

lista= ['8.8.8.7', '4.4.4.4', '8.8.4.4','192.168.1.82','8.8.8.6','10.0.0.0']
contador=1
for ipAddress in lista:
    datos=ping(ipAddress)
  
    if datos != None:
         contador +=1
         book = openpyxl.load_workbook('ping.xlsx')
         sheet = book.active
         sheet[f'A{contador}']=ipAddress
         sheet[f'B{contador}']=datos
         book.save('ping.xlsx')
         print(f'Guardado! {ipAddress}')
         

    if datos == None:
        datos2= datos
        iplista=ipAddress.split(".")
        last=(int(iplista.pop()) + 1)
        iplista.append(last)
        a, b , c , d =iplista
        ipNueva = f'{a}.{b}.{c}.{str(d)}'
        datos2=ping(ipNueva)
      
    
    if datos2 != None :  
         contador +=1
         book = openpyxl.load_workbook('ping.xlsx')
         sheet = book.active
         sheet[f'A{contador}']=ipNueva
         sheet[f'B{contador}']=datos2
         book.save('ping.xlsx')
         print(f'Guardado ip nueva! {ipNueva}')
    
    if datos == None:
        datos3= datos
        iplista=ipAddress.split(".")
        last=(int(iplista.pop()) + 2)
        iplista.append(last)
        a, b , c , d =iplista
        ipNueva2= f'{a}.{b}.{c}.{str(d)}'
        datos3=ping(ipNueva2)
        
    if datos3 != None :  
        contador +=1
        book = openpyxl.load_workbook('ping.xlsx')
        sheet = book.active
        sheet[f'A{contador}']=ipNueva2
        sheet[f'B{contador}']=datos3
        book.save('ping.xlsx')
        print(f'Guardado ip nueva! {ipNueva2}')    
       
   
    else:
        print("sin Datos")
   

